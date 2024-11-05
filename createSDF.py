import pandas as pd
import openpyxl as op
import sys


input_tab = 'Input Data'
default_tab ='Default values for Trueview'
campaign_name='Campaign Name'
start_date='Start Date'
end_date='End Date'
included_locations='Included Locations'
excluded_locations='Excluded Locations'
budget='Budget'
creative_type='Creative Type'
advertiser = 'Advertiser'
name = 'Name'
status = 'Status'
campaign_id='Campaign Id'
io_id='Io Id'
line_item_id='Line Item Id'
ad_group_id='Ad Group Id'
ad_id='Ad Id'
campaign_goal = 'Campaign Goal'
campaign_goal_kpi = 'Campaign Goal KPI'
campaign_goal_kpi_value = 'Campaign Goal KPI Value'
frequency_enabled = 'Frequency Enabled'
frequency_amount = 'Frequency Amount'
frequency_exposures = 'Frequency Exposures'
campaign_start_date = 'Campaign Start Date'
campaign_end_date = 'Campaign End Date'
timestamp = 'Timestamp'
type = 'Type'
io_name = 'Io Name'
budget_type = 'Budget Type'
line_item_name = 'Line Item Name'
targeting_genders = 'Targeting Genders'
targeting_age_ranges = 'Targeting Age Ranges'
video_ad_format = 'Video Ad Format'
bid_cost = 'Bid Cost'
bid_amount = 'Bid Amount'
ad_group_name = 'Ad Group Name'
display_url = 'Display Url'
landing_page_url = 'Landing Page URL'
action_button_label = 'Action Button Label'
call_to_action = 'Call To Action'
frequency='Frequency'


# Function to read input file and returns the default values and input data
def read_inputfile(input_data):
    """
    Returns a list of default values, row data and campaign names after taking the input sheet as input
    :param input_data: A string which contains the filename of the input file.
    :return default_data : A dict containing the default values in a header:value format
    :return row_data : A list of dicts containing the header and value where each
    :return campaigns : A list containing the name of campaigns
    """
    input_sheet = input_data[input_tab]
    default_sheet_trueview = input_data[default_tab]
    default_data = {header.value: value.value for (header, value) in
                    zip(default_sheet_trueview[1], default_sheet_trueview[2]) if header.value is not None}
    row_data_headers = [cell.value for cell in input_sheet[2] if cell.value is not None]
    campaigns=[]
    row_data = []

    try:
        for col in input_sheet.iter_cols(min_row=1, min_col=2,max_row=1):  # start from row 2 to skip the header
            for cell in col:
                if cell.value is not None:
                    campaigns.append(cell.value)

        for row in input_sheet.iter_rows(min_row=3, values_only=True):
            if all(cell is None for cell in row):
                break
            row_dict={}
            row_dict.update({row_data_headers[i]: row[i] for i in range(0, len(row_data_headers))})
            row_data.append(row_dict)

        # Caling function for validation
        validateRow(campaigns, row_data)

    except ValueError as e:
        print(e)
        sys.exit()

    return default_data, row_data,campaigns

def validateRow(campaigns,input):
    """
    Returns a ValueError Exception if the data does not meet the validation conditions
    :param campaigns : A list containing the names of campaigns
    :param input: A list of dicts that contains row data in a header, value format
    :return: ValueError if the data does not meet the validation conditions
    """
    non_empty_fields={campaign_name:campaigns,
                  start_date:[row[start_date] for row in input],
                  end_date:[i[end_date] for i in input],
                  included_locations:[i[included_locations] for i in input],
                  budget:[i[budget] for i in input],
                  creative_type:[i[creative_type] for i in input]}

    empty_list = [field for field,values in non_empty_fields.items() if
                  any(value is None for value in values)or not values]
    if empty_list:
        if len(empty_list)==1:
            raise ValueError(f'Field {"".join(empty_list)} is Mandatory and cannot be empty!')
        else:
            raise ValueError(f'Fields {", ".join(empty_list[:-1])} and {empty_list[-1]} are Mandatory and cannot be empty!')


def generateSdfCampaign(default,rows,campaigns,id_counter):
    """
    Returns the campaign ids created and the id counter after creating the sdf-campaigns file
    :param default: A dict containing the default values as header, value pair
    :param rows: A list of dicts containing the data for each row as a header,value pair
    :param campaigns: A list containing the names of campaigns
    :param id_counter: An integer value that holds the value of the current id to be passed
    :return campaign_ids: A list containing the campaign ids
    :return id_counter: An integer value that holds the next id counter after creating sdf-campaigns file
    """
    sdf_campaign = []
    for index in range(len(campaigns)):
        entry = {
        campaign_id: f'ext{id_counter}',
        advertiser: f'1707036-{id_counter}',
        name: f'{campaigns[index]}',
        status : default[status],
        campaign_goal: default[campaign_goal],
        campaign_goal_kpi: default[campaign_goal_kpi],
        campaign_goal_kpi_value : default[campaign_goal_kpi_value],
        frequency_enabled: default[frequency_enabled],
        budget: sum(d[budget] for d in rows),
        frequency_exposures : default[frequency_amount],
        campaign_start_date: rows[0][start_date].strftime('%d/%m/%Y'),
        campaign_end_date : rows[0][end_date].strftime('%d/%m/%Y')
        }
        sdf_campaign.append(entry)
        id_counter=id_counter+1
    pd.DataFrame(sdf_campaign).to_csv('/Users/gokul.raj/Downloads/test1/output/SDF-Campaigns.csv', index=False)
    # print(f'campaigns : {sdf_campaign}')
    campaign_ids=[item[campaign_id] for item in sdf_campaign]
    return campaign_ids,id_counter

def generateSdfInsertionOrders(default,rows,campaign_ids,campaign_names,id_counter):
    """
    Returns the insertion order names ids and the id counter after creating sdf-insertionorders file
    :param default: a dict containing default values as header, value format
    :param rows: A list of dicts containing the data for each row as a header,value pair
    :param campaign_ids: A list containing campaign_ids
    :param campaign_names: A list containing campaign names
    :param id_counter: An integer holding the value of the next id
    :return io_names: A list containing the insertion order names
    :return io_ids: A list containing the insertion order ids
    :return id_counter: An integer containing the value of the next id
    """
    sdf_insertionorders=[]
    for outer_index in range(len(campaign_ids)):
        for inner_index in range(len(rows)):
            entry = {
            io_id: f'ext{id_counter}',
            campaign_id: campaign_ids[outer_index],
            name: f'{rows[inner_index][creative_type].split("-")[1]}__{rows[inner_index][included_locations]}_{campaign_names[outer_index]}',
            timestamp : rows[inner_index].get(timestamp,None),
            frequency_enabled : default[frequency_enabled]
            }
            sdf_insertionorders.append(entry)
            id_counter=id_counter+1

    pd.DataFrame(sdf_insertionorders).to_csv(
        '/Users/gokul.raj/Downloads/test1/output/SDF-InsertionOrders.csv', index=False)
    # print(f'io : {sdf_insertionorders}')
    io_names = [item[name] for item in sdf_insertionorders]
    io_ids= [item[io_id] for item in sdf_insertionorders]
    return io_names,io_ids,id_counter

def generateSdfLineItems(default,rows,io_ids,io_names,id_counter,campaign_ids):
    """
    Returns the line item ids, names and the id counter after creating the sdf-lineitems file
    :param default: a dict containing default values as header, value format
    :param rows: A list of dicts containing the data for each row as a header,value pair
    :param io_ids: A list containing insertion order ids
    :param io_names: A list containing insertion order names
    :param id_counter: An integer that holds the value of next id
    :param campaign_ids: a list of campaign ids
    :return li_ids: A list containing line item ids
    :return li_names: A list containing line item names
    :return id_counter : An integer value that hold the next value of id
    """
    io_counter = 0
    sdf_lineitems=[]
    for outer_index in range(len(campaign_ids)):
        for inner_index in range(len(rows)):
            entry = {
                line_item_id : f'ext{id_counter}',
                io_id : io_ids[io_counter],
                io_name : io_names[io_counter],
                type : rows[inner_index].get(type, None),
                name : f'{io_names[io_counter].split("_")[2]}__{io_names[io_counter].split("__")[0]}',
                timestamp : rows[inner_index].get(timestamp, None),
                status : default[status],
                start_date : rows[inner_index][start_date].strftime('%d/%m/%Y %H:%M'),
                end_date : rows[inner_index][end_date].strftime('%d/%m/%Y %H:%M'),
                budget_type : 'TrueView Budget',
                frequency_exposures : rows[inner_index][frequency]
            }
            sdf_lineitems.append(entry)
            io_counter=io_counter+1
            id_counter=id_counter+1
    pd.DataFrame(sdf_lineitems).to_csv('/Users/gokul.raj/Downloads/test1/output/SDF-LineItems.csv', index=False)
    # print(f'li : {sdf_lineitems}')
    li_names = [item[name] for item in sdf_lineitems]
    li_ids = [item[line_item_id] for item in sdf_lineitems]
    return li_ids,li_names,id_counter

def generateSdfAdGroups(default,rows,li_ids,li_names,id_counter,campaign_ids):
    """
    Returns the Ad group ids, names and the id counter after creating the sdf-AdGroups file
    :param default: a dict containing default values as header, value format
    :param rows: A list of dicts containing the data for each row as a header,value pair
    :param li_ids: A list containing the line item ids
    :param li_names: A list containing line item names
    :param id_counter: An integer value that holds the value of the next id
    :param campaign_ids: A list containing campaign ids
    :return ag_ids: A list containing the Ad group ids
    :return ag_names: A list containing names of Ad groups
    :return id_counter : AN integer hold ing the value of the next id
    """
    li_counter = 0
    sdf_adgroups = []
    for outer_index in range(len(campaign_ids)):
        for inner_index in range(len(rows)):
            name_ad = f'{li_names[li_counter]}__dbm_{rows[inner_index][included_locations]}'
            if rows[inner_index][excluded_locations] is not None:
                name_ad = f'{name_ad}_Exclude_{rows[inner_index][excluded_locations]}'
            name_ad = f'{name_ad}_{rows[inner_index][targeting_genders]}_{rows[inner_index][targeting_age_ranges]}'
            entry={
                ad_group_id : f'ext{id_counter}',
                line_item_id : li_ids[li_counter],
                line_item_name : li_names[li_counter],
                name : name_ad,
                status : default[status],
                video_ad_format : rows[inner_index][creative_type].split('-')[1],
                bid_cost : rows[inner_index][bid_amount]

            }
            sdf_adgroups.append(entry)
            li_counter=li_counter+1
            id_counter=id_counter+1

    pd.DataFrame(sdf_adgroups).to_csv('/Users/gokul.raj/Downloads/test1/output/SDF-AdGroups.csv', index=False)
    # print(f'ag : {sdf_adgroups}')
    ag_names = [item[name] for item in sdf_adgroups]
    ag_ids = [item[ad_group_id] for item in sdf_adgroups]
    return ag_ids, ag_names, id_counter

def generateSdfAdGroupAds(default,rows,ag_ids,ag_names,id_counter,campaign_ids):
    """
    Returns the Ad group ad ids, names and the id counter after creating the sdf-AdGroupsAds file
    :param default: a dict containing default values as header, value format
    :param rows: A list of dicts containing the data for each row as a header,value pair
    :param ag_ids: A list containing ad group ids
    :param ag_names: A list containing ad group names
    :param id_counter: An integer value which holds the next id
    :param campaign_ids: A list containing the campaign ids
    """
    ag_counter = 0
    sdf_adgroupads=[]
    for outer_index in range(len(campaign_ids)):
        for inner_index in range(len(rows)):
            name_parts = (ag_names[ag_counter].split('_', 3))
            entry ={
                ad_id : f'ext{id_counter}',
                ad_group_id : ag_ids[ag_counter],
                ad_group_name : ag_names[ag_counter],
                name : f'{name_parts[0]}{name_parts[-1]}',
                status : default[status],
                display_url : rows[inner_index][display_url],
                landing_page_url : rows[inner_index][landing_page_url],
                action_button_label : rows[inner_index][call_to_action]
            }
            sdf_adgroupads.append(entry)
            ag_counter=ag_counter+1
            id_counter=id_counter+1
    pd.DataFrame(sdf_adgroupads).to_csv('/Users/gokul.raj/Downloads/test1/output/SDF-AdGroupAds.csv', index=False)
    # print(f'agad : {sdf_adgroupads}')



# Main Function
if __name__ == '__main__':
    try:
        input_file = op.load_workbook(f'/Users/gokul.raj/Downloads/test1/ad_builder_dbm_media_upload_template.xlsx')
    except FileNotFoundError as e:
        print('Input File not found')
        sys.exit()

    # Function Calls
    id_counter=1
    default, rows, campaigns = read_inputfile(input_file)
    c_ids,id_counter=generateSdfCampaign(default,rows,campaigns,id_counter)
    io_names,io_ids,id_counter=generateSdfInsertionOrders(default,rows,c_ids,campaigns,id_counter)
    li_ids,li_names,id_counter=generateSdfLineItems(default,rows,io_ids,io_names,id_counter,c_ids)
    ag_ids,ag_names,id_counter=generateSdfAdGroups(default,rows,li_ids,li_names,id_counter,c_ids)
    generateSdfAdGroupAds(default,rows,ag_ids, ag_names, id_counter,c_ids)



